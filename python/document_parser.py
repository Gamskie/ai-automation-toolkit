"""
Excel/Word template parser with prefer_below logic.

Real-world use case: parsing IPL (Instruction to Load) templates from a
commodity exporter. The labels and values aren't always in the same row —
sometimes the value sits BELOW the label, sometimes to the right. A naive
"value to the right of label" parser breaks. This module handles both.

Author: Gamma Wira <gammawirawibowo@gmail.com>
License: MIT
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Optional

import openpyxl  # pip install openpyxl


@dataclass
class CellMatch:
    sheet: str
    row: int
    col: int
    value: str


def find_label_value(
    workbook_path: str,
    label: str,
    *,
    prefer_below: bool = False,
    case_insensitive: bool = True,
    max_distance: int = 3,
) -> Optional[CellMatch]:
    """Find a label cell and return the value cell next to/under it.

    Args:
        workbook_path: path to .xlsx
        label: text to search for, e.g. "POD" or "Vessel Name"
        prefer_below: if True, look for the value DIRECTLY UNDER the label first.
                      Falls back to looking RIGHT if nothing is below.
                      If False (default), looks right first, then below.
        max_distance: how many cells to scan in each direction.

    Returns:
        CellMatch with sheet, row, col, and string value — or None.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    needle = label.lower() if case_insensitive else label

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                cmp_text = cell_text.lower() if case_insensitive else cell_text
                if needle not in cmp_text:
                    continue

                # Found a label-ish cell. Now hunt for its value.
                directions = (
                    [(1, 0), (0, 1)] if prefer_below else [(0, 1), (1, 0)]
                )
                for dr, dc in directions:
                    for step in range(1, max_distance + 1):
                        nr = cell.row + dr * step
                        nc = cell.column + dc * step
                        try:
                            nv = ws.cell(row=nr, column=nc).value
                        except IndexError:
                            break
                        if nv is None or str(nv).strip() == "":
                            continue
                        return CellMatch(
                            sheet=sheet,
                            row=nr,
                            col=nc,
                            value=str(nv).strip(),
                        )
    return None


# ---- Convenience: pull a known set of fields from an IPL template ----

IPL_FIELDS = {
    "vessel_name":  {"label": "Vessel Name", "prefer_below": False},
    "voyage":       {"label": "Voyage",      "prefer_below": False},
    "pol":          {"label": "POL",         "prefer_below": False},
    "pod":          {"label": "POD",         "prefer_below": True},   # value sits below in this template
    "etd":          {"label": "ETD",         "prefer_below": False},
    "buyer":        {"label": "Buyer",       "prefer_below": False},
    "commodity":    {"label": "Commodity",   "prefer_below": False},
    "qty_ctn":      {"label": "Qty Container", "prefer_below": False},
}


def parse_ipl(path: str) -> dict[str, Optional[str]]:
    """Parse an IPL template and return a flat dict of known fields."""
    out: dict[str, Optional[str]] = {}
    for key, spec in IPL_FIELDS.items():
        match = find_label_value(path, spec["label"], prefer_below=spec["prefer_below"])
        out[key] = match.value if match else None
    return out


# ---- demo ------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    import json

    ap = argparse.ArgumentParser(description="Parse an IPL .xlsx and print known fields as JSON.")
    ap.add_argument("path", help="Path to IPL .xlsx file")
    args = ap.parse_args()

    print(json.dumps(parse_ipl(args.path), indent=2, ensure_ascii=False))
